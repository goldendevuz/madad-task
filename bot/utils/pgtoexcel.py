import openpyxl
from openpyxl.styles import Font


async def export_to_excel(data,filepath):

    headings = list(data[0].keys())
    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.row_dimensions[1].font = Font(bold=True)

    # Spreadsheet row and column indexes start at 1,
    # so we use "start = 1" in enumerate, so
    # we don't need to add 1 to the indexes.
    for col_number, heading in enumerate(headings, start=1):
        sheet.cell(row=1, column=col_number).value = heading

    # This time we use "start = 2" to skip the heading row.
    for row_number, rowdata in enumerate(data, start=2):
        for col_number,(key,value) in enumerate(rowdata.items(), start=1):
            sheet.cell(row=row_number, column=col_number).value = value

    wb.save(filepath)